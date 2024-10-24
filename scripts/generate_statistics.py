import os
import sys

import numpy as np
import openpyxl
import pandas as pd
from scipy.stats import linregress

from constants import EXERCISES

sys.path.append(os.path.join(os.path.dirname(__file__), ".."))

from markerless_mocap_validation_for_team_sports.statistics.bland_altmann import \
    bland_altmann_statistics  # noqa

JOINTS: list[str] = [
    "shoulder",
    "elbow",
    "wrist",
    "hip",
    "knee",
    "ankle",
]

SEGMENTS: list[str] = [
    "head",
    "torso",
    "upper arm",
    "lower arm",
    "hand",
    "pelvis",
    "thigh",
    "shank",
    "foot",
]

ANGLES_TO_DROP = [
    "pelvis[0]",
    "pelvis[1]",
    "pelvis[2]",
    "thorax[0]",
    "thorax[1]",
    "thorax[2]",
    "head[0]",
    "head[1]",
    "head[2]",
    "left_elbow[1]",
    "left_wrist[2]",
    "right_elbow[1]",
    "right_wrist[2]",
]
POINTS_TO_DROP = [
    "pelvis[3]",
    "left_thigh[3]",
    "left_shank[3]",
    "left_foot[3]",
    "left_toes[3]",
    "right_thigh[3]",
    "right_shank[3]",
    "right_foot[3]",
    "right_toes[3]",
    "head[3]",
    "torso[3]",
    "left_upper_arm[3]",
    "left_lower_arm[3]",
    "left_hand[3]",
    "right_upper_arm[3]",
    "right_lower_arm[3]",
    "right_hand[3]",
    "centre_of_mass[3]",
    "lab[3]",
]
ROTATIONS_TO_DROP = [
    "left_thigh[3]",
    "left_shank[3]",
    "left_foot[3]",
    "left_toes[3]",
    "left_upper_arm[3]",
    "left_lower_arm[3]",
    "left_hand[3]",
    "pelvis[3]",
    "torso[3]",
    "head[3]",
    "right_thigh[3]",
    "right_shank[3]",
    "right_foot[3]",
    "right_toes[3]",
    "right_upper_arm[3]",
    "right_lower_arm[3]",
    "right_hand[3]",
    "lab[3]",
]


def main(input_dir: str, exercise: str, processing_info: str = "preprocessing_df.csv"):
    """Function to generate statistics for a given exercise.
    For every exercise, the output can be sligthly different. For instance, for
    sidesteps it should get all statistics for the sidestep side and non sidestep
    side, but which individual does not really matter. Contrary, for the dribble,
    there is not really a side of interest, so th excel file will be different.

    The final output will be a excel file with the following statistics:
    - sTEE
    - r
    - RMSD
    - Bland altmann (bias + 95% confidence interval)

    This will be calculated for all joints and segments, and for the CoM in x, y and z.

    Args:
        input_dir (str): The directory where the data is stored (.parquet files).
        exercise (str): The exercise for which the statistics should be calculated
        processing_info (str, optional): The preprocessing info.
            Defaults to "preprocessing_df.csv".

    Returns:
        None
    """

    if exercise not in EXERCISES:
        raise ValueError(f"Exercise should be one of {EXERCISES}, not {exercise}")
    if not os.path.exists(input_dir):
        raise FileNotFoundError(f"Directory {input_dir} does not exist")

    preprocessing_df = pd.read_csv(processing_info, index_col=0)

    all_points_vicon = []
    all_points_theia = []
    all_angles_vicon = []
    all_angles_theia = []
    all_rotations_vicon = []
    all_rotations_theia = []

    right_list_map = {
        "points": [all_points_vicon, all_points_theia],
        "angles": [all_angles_vicon, all_angles_theia],
        "rotations": [all_rotations_vicon, all_rotations_theia],
    }

    vicon_files = [
        os.path.join(input_dir, dir_, file)
        for dir_ in os.listdir(input_dir)
        if os.path.isdir(os.path.join(input_dir, dir_))
        for file in os.listdir(os.path.join(input_dir, dir_))
        if file.endswith(".parquet") and exercise in file and "vicon" in file
    ]
    if exercise == "sidestep":  # distinguis between sidestep and double sidestep
        vicon_files = [x for x in vicon_files if "double" not in x]

    for file in vicon_files:
        right_key = (
            "points"
            if "points" in file
            else "angles" if "angles" in file else "rotations"
        )
        vicon_df = pd.read_parquet(file).reset_index(drop=False, names="time")
        theia_df = pd.read_parquet(file.replace("vicon", "theia")).reset_index(
            drop=False, names="time"
        )

        particpant_number = file.split("/")[-2]
        base_name = file.split("/")[-1]
        trial_number = max([int(base_name[x]) for x in get_digit_idxs(base_name)])
        current_processing_info = preprocessing_df[
            (preprocessing_df["participant_number"] == particpant_number)
            & (preprocessing_df["vicon_file"].str.contains(exercise))
            & (preprocessing_df["trial_number"] == trial_number)
        ].iloc[0]
        special_side = current_processing_info["special_side"]
        non_special_side = "right" if special_side == "left" else "left"

        # check for gimball lock
        if "angles" in file:
            for col in vicon_df.columns:
                if (vicon_df[col] == 1).all():
                    vicon_df[col] = np.nan
                if (theia_df[col] == 1).all():
                    theia_df[col] = np.nan
            # Remove unnecessary columns
            vicon_df["left_elbow[2]"] = vicon_df["left_wrist[2]"]
            vicon_df["right_elbow[2]"] = vicon_df["right_wrist[2]"]

            vicon_df = vicon_df.drop(columns=ANGLES_TO_DROP)
            theia_df = theia_df.drop(columns=ANGLES_TO_DROP)

        elif "points" in file:
            vicon_df = vicon_df.drop(columns=POINTS_TO_DROP)
            theia_df = theia_df.drop(columns=POINTS_TO_DROP)

        elif "rotations" in file:
            vicon_df = vicon_df.drop(columns=ROTATIONS_TO_DROP)
            theia_df = theia_df.drop(columns=ROTATIONS_TO_DROP)

        if not pd.isnull(special_side):
            rename = {
                col: col.replace(special_side, "special_side")
                for col in vicon_df.columns
                if special_side in col
            }
            rename.update(
                {
                    col: col.replace(non_special_side, "non_special_side")
                    for col in vicon_df.columns
                    if non_special_side in col
                }
            )
            vicon_df = vicon_df.rename(columns=rename)
            theia_df = theia_df.rename(columns=rename)

        vicon_df["participant_number"] = particpant_number
        theia_df["participant_number"] = particpant_number
        right_list_map[right_key][0].append(vicon_df)
        right_list_map[right_key][1].append(theia_df)

    combined_data = {
        "points_vicon_df": pd.concat(all_points_vicon, ignore_index=True),
        "points_theia_df": pd.concat(all_points_theia, ignore_index=True),
        "angles_vicon_df": pd.concat(all_angles_vicon, ignore_index=True),
        "angles_theia_df": pd.concat(all_angles_theia, ignore_index=True),
        "rotations_vicon_df": pd.concat(all_rotations_vicon, ignore_index=True),
        "rotations_theia_df": pd.concat(all_rotations_theia, ignore_index=True),
    }

    columns = [
        [name + " x", name + " y", name + " z"]
        for name in ["sTEE", "r", "RMSD", "BA Bias", "BA CI"]
    ]
    if exercise in ["double sidestep", "sidestep"]:
        index = [
            "Head Orientation",
            "Torso Orientation",
            "Upper arm Orientation sidestep side",
            "Upper arm Orientation non sidestep side",
            "Lower arm Orientation sidestep side",
            "Lower arm Orientation non sidestep side",
            "Hand Orientation sidestep side",
            "Hand Orientation non sidestep side",
            "Pelvis Orientation",
            "Thigh Orientation sidestep side",
            "Thigh Orientation non sidestep side",
            "Shank Orientation sidestep side",
            "Shank Orientation non sidestep side",
            "Foot Orientation sidestep side",
            "Foot Orientation non sidestep side",
            "Shoulder angle sidestep side",
            "Shoulder angle non sidestep side",
            "Elbow angle sidestep side",
            "Elbow angle non sidestep side",
            "Wrist angle sidestep side",
            "Wrist angle non sidestep side",
            "Hip angle sidestep side",
            "Hip angle non sidestep side",
            "Knee angle sidestep side",
            "Knee angle non sidestep side",
            "Ankle angle sidestep side",
            "Ankle angle non sidestep side",
            "Centre of Mass",
        ]

        df = pd.DataFrame(
            index=index, columns=[name for sublist in columns for name in sublist]
        )

        stats = save_statistics(combined_data, df, "sidestep")

    elif exercise == "shot":
        index = [
            "Head Orientation",
            "Torso Orientation",
            "Upper arm Orientation",
            "Lower arm Orientation",
            "Hand Orientation",
            "Pelvis Orientation",
            "Thigh Orientation shot side",
            "Thigh Orientation non shot side",
            "Shank Orientation shot side",
            "Shank Orientation non shot side",
            "Foot Orientation shot side",
            "Foot Orientation non shot side",
            "Shoulder angle",
            "Elbow angle",
            "Wrist angle",
            "Hip angle shot side",
            "Hip angle non shot side",
            "Knee angle shot side",
            "Knee angle non shot side",
            "Ankle angle shot side",
            "Ankle angle non shot side",
            "Centre of Mass",
        ]
        df = pd.DataFrame(
            index=index, columns=[name for sublist in columns for name in sublist]
        )
        stats = save_statistics(combined_data, df, "shot")

    elif exercise in ["dribble simulation", "slow dribble", "pass", "walk rotation"]:
        index = [
            "Head Orientation",
            "Torso Orientation",
            "Upper arm Orientation",
            "Lower arm Orientation",
            "Hand Orientation",
            "Pelvis Orientation",
            "Thigh Orientation",
            "Shank Orientation",
            "Foot Orientation",
            "Shoulder angle",
            "Elbow angle",
            "Wrist angle",
            "Hip angle",
            "Knee angle",
            "Ankle angle",
            "Centre of Mass",
        ]
        df = pd.DataFrame(
            index=index, columns=[name for sublist in columns for name in sublist]
        )
        stats = save_statistics(combined_data, df, exercise)
    elif exercise == "high five":
        index = [
            "Head Orientation",
            "Torso Orientation",
            "Upper arm Orientation high five side",
            "Upper arm Orientation non high five side",
            "Lower arm Orientation high five side",
            "Lower arm Orientation non high five side",
            "Hand Orientation high five side",
            "Hand Orientation non high five side",
            "Pelvis Orientation",
            "Thigh Orientation",
            "Shank Orientation",
            "Foot Orientation",
            "Shoulder angle high five side",
            "Shoulder angle non high five side",
            "Elbow angle high five side",
            "Elbow angle non high five side",
            "Wrist angle high five side",
            "Wrist angle non high five side",
            "Hip angle",
            "Knee angle",
            "Ankle angle",
            "Centre of Mass",
        ]
        df = pd.DataFrame(
            index=index, columns=[name for sublist in columns for name in sublist]
        )
        stats = save_statistics(combined_data, df, "high five")
    return stats


def get_digit_idxs(string: str):
    return [i for i, c in enumerate(string) if c.isdigit()]


def save_statistics(combined_data: dict, df: pd.DataFrame, special_side: str):
    """Function to calculate and save the statistics for the high five exercise.
    The statistics will be saved in a excel file with the following statistics:
    - sTEE
    - r
    - RMSD
    - Bland altmann (bias + range)

    This will be calculated for all joints and segments, and for the CoM in x, y and z.
    For the lower extremities, there will be no distinction between left and right.
    For the upper extremities, there will be a distinction between the high five side
    and the non high five side.

    Args:
        combined_data (dict): A dictionary with the following keys:
            - points_vicon_df: The dataframe with the points from the vicon system
            - points_theia_df: The dataframe with the points from the theia system
            - angles_vicon_df: The dataframe with the angles from the vicon system
            - angles_theia_df: The dataframe with the angles from the theia system
            - rotations_vicon_df: The dataframe with the rotations from the vicon system
            - rotations_theia_df: The dataframe with the rotations from the theia system
        prepocessing_info (str): The preprocessing info.

    Returns:
        None
    """

    for index_name in df.index:
        df_type = (
            "rotations"
            if "Orientation" in index_name
            else "angles" if "angle" in index_name else "points"
        )
        vicon_df = combined_data[f"{df_type}_vicon_df"]
        theia_df = combined_data[f"{df_type}_theia_df"]
        body_part = index_name.split(" ")[0].lower()

        # get the right selection of samples based on the special side
        if f"non {special_side}" in index_name:
            cols = [col for col in vicon_df.columns if "non_special_side" in col]
        elif special_side in index_name:
            cols = [
                col
                for col in vicon_df.columns
                if "special_side" in col and "non" not in col
            ]
        else:
            cols = vicon_df.columns

        cols = [col for col in cols if body_part in col]
        vicon_df = vicon_df[cols]
        theia_df = theia_df[cols]

        if len(vicon_df.columns) > 3:  # no special side, vertical stack right axis
            if body_part in ["upper", "lower"]:
                body_part += " arm"
            new_col_names = [f"{body_part}[{i}]" for i in range(3)]
            new_df_vicon = pd.DataFrame(
                index=range(vicon_df.shape[0] * 2), columns=new_col_names
            )
            new_df_theia = pd.DataFrame(
                index=range(vicon_df.shape[0] * 2), columns=new_col_names
            )
            for i in range(3):
                cols = [col for col in vicon_df.columns if f"[{i}]" in col]
                if len(cols) == 0:
                    continue

                new_df_vicon.loc[: vicon_df.shape[0] - 1, f"{body_part}[{i}]"] = (
                    vicon_df[cols[0]].values
                )
                new_df_vicon.loc[vicon_df.shape[0] :, f"{body_part}[{i}]"] = vicon_df[
                    cols[1]
                ].values
                new_df_theia.loc[: theia_df.shape[0] - 1, f"{body_part}[{i}]"] = (
                    theia_df[cols[0]].values
                )
                new_df_theia.loc[theia_df.shape[0] :, f"{body_part}[{i}]"] = theia_df[
                    cols[1]
                ].values
            vicon_df = new_df_vicon
            theia_df = new_df_theia

        statistics = calculate_statistics(
            vicon_df,
            theia_df,
            combined_data[f"{df_type}_vicon_df"]["participant_number"].values,
        )
        df.loc[index_name] = statistics

    beautify_to_excel(df, special_side)
    return df


def beautify_to_excel(df: pd.DataFrame, side_name: str):
    """Function to create the excel file with the statistics."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    columns = ["sTEE", "r", "RMSD", "BA Bias (CI)"]
    subcolumns = ["x", "y", "z"]
    start_column = 3
    start_row = 1
    for i, column in enumerate(columns):
        sheet.cell(row=start_row, column=start_column + i * 3, value=column)
        for j, subcolumn in enumerate(subcolumns):
            sheet.cell(
                row=start_row + 1, column=start_column + i * 3 + j, value=subcolumn
            )

    row_names = []
    subnames = {}
    for name in df.index:
        basename = name.split(f" non {side_name}")[0].split(f" {side_name}")[0]
        if basename not in row_names:
            row_names.append(basename)
            if side_name in name and "non" not in name:
                subnames[basename] = [side_name, f"non {side_name}"]

    current_row = 3
    base_column = 1
    for row_name in row_names:
        sheet.cell(row=current_row, column=base_column, value=row_name)
        if row_name in subnames:
            for subname in subnames[row_name]:
                sheet.cell(row=current_row, column=base_column + 1, value=subname)
                current_row += 1
        else:
            current_row += 1

    # stee values
    base_values = df.iloc[:, :3]
    start_row = 3
    start_column = 3
    for i in range(base_values.shape[0]):
        for j in range(base_values.shape[1]):
            value = base_values.iloc[i, j]
            cell = sheet.cell(
                row=start_row + i,
                column=start_column + j,
                value=f"{value:.2f}",
            )
            cell.number_format = "0.00"
            if value <= 0.2:
                cell.font = openpyxl.styles.Font(color="008000")
            elif value > 0.6:
                cell.font = openpyxl.styles.Font(color="FF7F00")

    # r values
    base_values = df.iloc[:, 3:6]
    start_row = 3
    start_column = 6
    for i in range(base_values.shape[0]):
        for j in range(base_values.shape[1]):
            value = base_values.iloc[i, j]
            cell = sheet.cell(
                row=start_row + i,
                column=start_column + j,
                value=f"{value:.2f}",
            )
            cell.number_format = "0.00"
            if value >= 0.9:
                cell.font = openpyxl.styles.Font(color="008000")
            elif value < 0.7:
                cell.font = openpyxl.styles.Font(color="FF7F00")

    # rmsd values
    base_values = df.iloc[:, 6:9]
    start_row = 3
    start_column = 9
    for i in range(base_values.shape[0]):
        for j in range(base_values.shape[1]):
            value = base_values.iloc[i, j]
            cell = sheet.cell(
                row=start_row + i,
                column=start_column + j,
                value=f"{value:.2f}",
            )
            cell.number_format = "0.00"
            if value <= 5:
                cell.font = openpyxl.styles.Font(color="008000")
            elif value > 12.5:
                cell.font = openpyxl.styles.Font(color="FF7F00")

    start_row = 3
    start_column = 12
    ba_values = df.iloc[:, 9:]
    for i in range(ba_values.shape[0]):
        for j in range(3):
            bias = ba_values.iloc[i, j]
            ci = ba_values.iloc[i, j + 3]
            cell = sheet.cell(
                row=start_row + i,
                column=start_column + j,
                value=f"{bias:.2f} ({ci:.1f})",
            )
            if ci <= 5:
                cell.font = openpyxl.styles.Font(color="008000")
            elif ci > 12.5:
                cell.font = openpyxl.styles.Font(color="FF7F00")

    workbook.save(f"{side_name} statistics.xlsx")


def calculate_statistics(
    vicon_df: pd.DataFrame, theia_df: pd.DataFrame, participant_numbers: np.ndarray
):
    """Function to calculate the statistics for a given dataframe. The statistics
    that will be calculated are:
    - sTEE
    - r
    - RMSD
    - Bland altmann (bias + range)

    Args:
        vicon_df (pd.DataFrame): The dataframe with the data from the vicon system.
        theia_df (pd.DataFrame): The dataframe with the data from the theia system.
        participant_numbers (np.ndarray): The participant numbers.

    Returns:
        dict: A dictionary with the following keys:
            - sTEE XYZ
            - r XYZ
            - RMSD XYZ
            - BA Bias (CI) XYZ
    """
    stee = np.zeros(3)
    r_values = np.zeros(3)
    rmsd = np.zeros(3)
    ba_bias = np.zeros(3)
    ba_ci = np.zeros(3)

    for i in range(3):
        current_col = [col for col in vicon_df.columns if f"[{i}]" in col]
        if len(current_col) == 0:
            continue
        not_nan_idxs = np.where(
            ~pd.isnull(vicon_df[current_col]) & ~pd.isnull(theia_df[current_col])
        )[0]
        vicon_values = (
            vicon_df.iloc[not_nan_idxs][current_col].values[:, 0].astype(float)
        )
        theia_values = (
            theia_df.iloc[not_nan_idxs][current_col].values[:, 0].astype(float)
        )

        if len(vicon_values) == 0:
            continue
        mask = (vicon_values < 0) | (theia_values < 0)
        other_option_vicon = np.where(mask, vicon_values + 360, vicon_values)
        other_option_theia = np.where(mask, theia_values + 360, theia_values)

        if (
            other_option_vicon.max() - other_option_vicon.min()
            < vicon_values.max() - vicon_values.min()
        ):
            vicon_values = other_option_vicon
            theia_values = other_option_theia

        _, _, r_value, _, _ = linregress(vicon_values, theia_values)
        stee[i] = np.sqrt((1 / (r_value**2)) - 1)
        r_values[i] = r_value
        diffs = vicon_values - theia_values
        rmsd[i] = np.sqrt(np.mean(diffs**2))
        bias, uloa, lloa = bland_altmann_statistics(vicon_values, theia_values)
        ba_bias[i] = bias
        ba_ci[i] = (uloa - lloa) / 2

    index = [
        [name + " x", name + " y", name + " z"]
        for name in ["sTEE", "r", "RMSD", "BA Bias", "BA CI"]
    ]
    res = pd.Series(
        np.concatenate([stee, r_values, rmsd, ba_bias, ba_ci]),
        index=[name for sublist in index for name in sublist],
    )
    return res


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--input_dir", type=str, default="data/preprocessed_dataset")
    parser.add_argument("--exercise", type=str, default="all")
    parser.add_argument(
        "--preprocessing_file_loc",
        type=str,
        default="preprocessing_df.csv",
    )

    args = parser.parse_args()
    if args.exercise == "all":
        for exercise in EXERCISES:
            main(args.input_dir, exercise, args.preprocessing_file_loc)
    else:
        main(args.input_dir, args.exercise, args.preprocessing_file_loc)
